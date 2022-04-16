import datetime
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = datetime.datetime.today()
ws['A2'] = '=sum(1, 2, 3)'
ws['A3'] = '=average(1, 2, 3)'

ws['A4'] = 10
ws['A5'] = 20
ws['A6'] = '=sum(A4:A5)'

ws.column_dimensions['A'].width = 30

wb.save('sample_formula.xlsx')
