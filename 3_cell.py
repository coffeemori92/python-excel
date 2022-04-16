from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = 'NadoSheet'

# A1셀에 1이라는 값을 입력
ws['A1'] = 1
ws['A2'] = 2
ws['A3'] = 3

ws['B1'] = 1
ws['B2'] = 2
ws['B3'] = 3

print(ws['A1']) # 셀 정보 출력
print(ws['A1'].value) # 셀의 값 출력

print(ws.cell(1, 1).value) # ws['A1'].value

# 반복문을 이용해서 데이터 입력
for i in range(0, 10):
  for j in range(0, 10):
    ws.cell(i + 1, j + 1, randint(0, 100))

wb.save('sample.xlsx')
