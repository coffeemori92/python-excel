from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet() # 새로운 시트 기본이름으로 생성
ws.title = 'MySheet'
ws.sheet_properties.tabColor = 'ff66ff' # 탭 색상 변경

ws1 = wb.create_sheet('YourSheet') #  주어진 이름으로 새로운 시트 생성
ws2 = wb.create_sheet('NewSheet', 2) # 2번째 index에 Sheet 생성

new_ws = wb['NewSheet']

# Sheet 복사
new_ws['A1'] = 'TEST' # A1에 값 대입
target = wb.copy_worksheet(new_ws)
target.title = 'Copied Sheet'

print(wb.sheetnames) # 모든 Sheet 이름 확인

wb.save('sample.xlsx')
