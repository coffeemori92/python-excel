from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

# 한줄씩 데이터 넣기
ws.append(['번호', '영어', '수학'])

for i in range(0, 10):
  ws.append([i + 1, randint(0, 100), randint(0, 100)])

col_B = ws['B'] # B 칼럼만 가지고 오기
print(col_B)

for cell in col_B:
  print(cell.value)

col_range = ws['B:C'] # 영어, 수학 칼럼 함께 가지고 오기
print(col_range)

for cols in col_range:
  for cell in cols:
    print(cell.value)

row_title = ws[1] # 1번째 row만 가지고 오기
print(row_title)

for cell in row_title:
  print(cell.value)

row_range = ws[2:6] # 2번째 줄에서 6번째 줄까지 가지고 오기
for rows in row_range:
  for cell in rows:
    print(cell.value, end = ' ')
  print()  

print() 

row_range = ws[2:ws.max_row] # 2번째 줄에서 마지막 줄까지 가지고 오기
for rows in row_range:
  for cell in rows:
    # print(cell.value, end = ' ')
    # print(cell.coordinate, end = ' ')
    xy = coordinate_from_string(cell.coordinate)
    # print(xy, end = ' ')
    print(xy[0], end = '')
    print(xy[1], end = ' ')
  print() 

print()

print(tuple(ws.rows))
for row in tuple(ws.rows):
  print(row[1].value)

print()

print(tuple(ws.columns))
for column in tuple(ws.columns):
  print(column[0].value)

for row in ws.iter_rows():
  print(row[1].value)

for column in ws.iter_cols():
  print(column[1].value)

print()

for row in ws.iter_rows(
  min_row = 2, max_row = 11, min_col = 2, max_col = 3
  ):
  print(row[0].value, row[1].value)  

wb.save('sample.xlsx')
