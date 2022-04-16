from openpyxl import load_workbook

wb = load_workbook('sample.xlsx') # 파일에서 워크북을 불러옴
ws = wb.active

for i in range(0, 10):
  for j in range(0, 10):
    print(ws.cell(i + 1, j + 1).value, end = ' ')
  print()

print()

for i in range(0, ws.max_row):
  for j in range(0, ws.max_column):
    print(ws.cell(i + 1, j + 1).value, end = ' ')
  print()

